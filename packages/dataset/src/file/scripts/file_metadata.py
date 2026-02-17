import sys
import os
import json
import csv

file_path = sys.argv[1]

info = {
    "file_name": os.path.basename(file_path),
    "extension": os.path.splitext(file_path)[1].lower(),
    "size_bytes": os.path.getsize(file_path) if os.path.exists(file_path) else None
}

if info["size_bytes"] is not None:
    units = ["B", "KB", "MB", "GB", "TB", "PB"]
    size = float(info["size_bytes"])
    idx = 0
    while size >= 1024 and idx < len(units) - 1:
        size /= 1024
        idx += 1
    info["size_human"] = f"{size:.2f} {units[idx]}"

row_estimate = None
column_estimate = None
header_preview = None

try:
    if info["extension"] in [".xlsx", ".xls"]:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        sheets = []
        first_sheet_rows = 0
        first_sheet_columns = 0
        for idx, ws in enumerate(wb.worksheets):
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            if idx == 0:
                first_sheet_rows = rows
                first_sheet_columns = cols
            sheet_info = {"name": ws.title, "rows": rows, "columns": cols}
            sheets.append(sheet_info)
        info["sheet_stats"] = sheets
        row_estimate = first_sheet_rows
        column_estimate = first_sheet_columns if first_sheet_columns > 0 else None
        try:
            first_sheet = wb.worksheets[0]
            header_preview = [str(cell.value) if cell.value is not None else "" for cell in next(first_sheet.iter_rows(min_row=1, max_row=1))]
        except Exception:
            header_preview = None
    else:
        row_count = 0
        with open(file_path, "rb") as fh:
            for _ in fh:
                row_count += 1
        row_estimate = row_count - 1 if row_count > 0 else 0
        with open(file_path, "r", encoding="utf-8", errors="ignore") as fh:
            reader = csv.reader(fh)
            header_preview = next(reader, [])
        column_estimate = len(header_preview) if header_preview else None
except Exception as error:
    info["metadata_error"] = str(error)

if row_estimate is not None:
    info["row_count_estimate"] = row_estimate
if column_estimate is not None:
    info["column_count_estimate"] = column_estimate
if header_preview is not None:
    info["header_preview"] = header_preview

print(json.dumps(info, ensure_ascii=False))

